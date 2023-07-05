# добавление библиотек
import tkinter
from tkinter import  ttk
from tkinter import *
from PIL import ImageTk, Image
import openpyxl
from datetime import datetime as date

name_date = date.today().strftime("%A")

book = openpyxl.load_workbook(filename='Книга1.xlsx')

if name_date == 'Monday':
    sheet = book['Monday']
elif name_date == 'Tuesday':
    sheet = book['Tuesday']
elif name_date == 'Wednesday':
    sheet = book['Wednesday']
elif name_date == 'Thursday':
    sheet = book['Thursday']
elif name_date == 'Friday':
    sheet = book['Friday']
else:
    sheet = book['Saturday']

m_row = sheet.max_row
m_column = sheet.max_column

def raise_frame(frame):
    frame.tkraise()

root = Tk()

root.title('Схема и расписание школы №22')

x = root.winfo_screenwidth()  # размер  по горизонтали монитора пользователя
y = root.winfo_screenheight() # размер по вертикали монитора пользователя
root.geometry('{}x{}'.format(int(x*1), int(y*1))) # размер приложения

f1 = Frame(root) # создание 1 рамки
f2 = Frame(root) # создание 2 рамки
f3 = Frame(root) # создание 3 рамки

# добавление заднего фона
image = Image.open("image.png") # открытие изображения
photo = ImageTk.PhotoImage(image.resize((1600, 600), Image.ANTIALIAS))
background_label1 = Label(f1, image=photo)
background_label1.pack(side="top", fill="both", expand="no")
background_label2 = Label(f2, image=photo)
background_label2.pack(side="top", fill="both", expand="no")
background_label3 = Label(f3, image=photo)
background_label3.pack(side="top", fill="both", expand="no")

mn={
    1:'08:00–08:40',
    2:'08:50–09:30',
    3:'09:45–10:25',
    4:'10:35–11:15',
    5:'11:35–12:15',
    6:'12:30–13:10',
    7:'13:20–14:00',
    8:'14:10–14:50',
    9:'15:00–15:40',
    10:'15:55–16:35',
    11:'16:45–17:25',
    12:'17:35–18:15',
    13:'18:20–19:00'
}

def search(b):
    newWindow = Toplevel(root)
    newWindow.title(f"Расписание {b}")
    newWindow.geometry('{}x{}'.format(int(x * 0.4), int(y * 0.4)))
    for i in range(2, m_row + 1):
        cell_obj = sheet.cell(row=i, column=1)
        if cell_obj.value == b:
            for j in range(1, m_column + 1):
                cell_object = sheet.cell(row=i, column=j + 1)
                if not (cell_object.value):
                    Label(newWindow, text=f"Уроков нет. {mn[j]}").pack()
                else:
                    Label(newWindow, text=f'{cell_object.value} {mn[j]}').pack()

def disableButton(my_button):
    my_button.config(state='disabled')

# функции с расписанием

def open102():
    b= 'Каб.102'
    search(b)
def open103():
    b= 'Каб.103'
    search(b)
def openInf1():
    b= 'каб. Инф.1'
    search(b)

def openInf2():
    b= 'каб. Инф.2'
    search(b)

def openHistory():
    b='Ист.'
    search(b)
def openRus2():
    b='каб.Рус2'
    search(b)
def openRus3():
    b='каб.Рус3'
    search(b)

def openRus4():
    b='каб.Рус4'
    search(b)

def openGym1():
    b='Зал1'
    search(b)

def openGym2():
    b='Зал2'
    search(b)

def openGeo():
    b= 'каб.Геог'
    search(b)

def openPhys1():
    b= 'Физика1'
    search(b)

def openPhys2():
    b= 'Физика2'
    search(b)

def openMath1():
    b= 'каб.Матем1'
    search(b)

def openMath2():
    b= 'каб.Матем2'
    search(b)

def openMath3():
    b= 'каб.Матем3'
    search(b)

def openChem():
    b= 'каб.Химии'
    search(b)

def openBio1():
    b= 'каб.Био1'
    search(b)

def openMath4():
    b= 'каб.Матем4'
    search(b)

def openART():
    b= 'ИЗО'
    search(b)

def openGERM():
    b= 'каб.Нем'
    search(b)

def openOBJ():
    b= 'ОБЖ'
    search(b)

def openTechM():
    b= 'Техн.М'
    search(b)

def openTechG():
    b= 'Техн.Д'
    search(b)

def openEngl2():
    b= 'Англ.2'
    search(b)

def openEngl1():
    b= 'Англ.1'
    search(b)

for frame in (f1, f2, f3):
    frame.grid(row=0, column=0, sticky='news')

# кнопки 1 этажа
btn_mov1 = Button(f1, text='Перейти на 2 этаж', command=lambda:raise_frame(f2))
btn_mov1.pack()
Label(f1, text='Этаж 1', font="Courier 30").pack()
btn_tim1 = Button(f1, text ="каб.103", command = open103)
btn_tim1.place(x=275, y=550)
btn_tim1 = Button(f1, text ="каб.102", command = open102)
btn_tim1.place(x=400, y=550)
btn_tim1 = Button(f1, text ="каб.Инф.1", command=openInf1)
btn_tim1.place(x=550, y=550)
btn_tim1 = Button(f1, text ="каб.Инф.2", command=openInf2)
btn_tim1.place(x=1100, y=550)
btn_tim1 = Button(f1, text ="каб.ИЗО", command=openART)
btn_tim1.place(x=1275, y=550)
btn_tim1 = Button(f1, text ="каб.Нем", command=openGERM)
btn_tim1.place(x=1425, y=500)
btn_tim1 = Button(f1, text ="каб.Технологии.Д", command=openTechM)
btn_tim1.place(x=1325, y=40)
btn_tim1 = Button(f1, text ="каб.Технологии.М", command=openTechG)
btn_tim1.place(x=475, y=40)
btn_tim1 = Button(f1, text ="каб.ОБЖ", command=openOBJ)
btn_tim1.place(x=225, y=40)
btn_tim1 = Button(f1, text ="столовая")
btn_tim1.place(x=1500, y=300)
btn_tim1.columnconfigure(1, weight=1)
btn_tim1.rowconfigure(1, weight=1)
btn_tim1.lift()

# кнопки 2 этажа
btn_mov2 = Button(f2, text='Перейти на 3 этаж', command=lambda:raise_frame(f3))
btn_mov2.pack()
Label(f2, text='Этаж 2', font="Courier 30").pack()
btn_tim2 = Button(f2, text ="каб.Ист", command=openHistory)
btn_tim2.place(x=550, y=550)
btn_tim2 = Button(f2, text ="каб.Р4", command=openRus4)
btn_tim2.place(x=125, y=350)
btn_tim2 = Button(f2, text ="каб.Р2", command=openRus2)
btn_tim2.place(x=400, y=550)
btn_tim2 = Button(f2, text ="каб.Р3", command=openRus3)
btn_tim2.place(x=275, y=550)
btn_tim2 = Button(f2, text ="спорт.Зал1", command=openGym1)
btn_tim2.place(x=475, y=300)
btn_tim2 = Button(f2, text ="каб.География", command=openGeo)
btn_tim2.place(x=1500, y=350)
btn_tim2 = Button(f2, text ="каб.Физики2", command=openPhys2)
btn_tim2.place(x=1100, y=550)
btn_tim2 = Button(f2, text ="каб.Физики1", command=openPhys1)
btn_tim2.place(x=1275, y=550)
btn_tim2 = Button(f2, text ="спорт.Зал2", command=openGym2)
btn_tim2.place(x=1000, y=300)
btn_tim2.columnconfigure(1, weight=1)
btn_tim2.rowconfigure(1, weight=1)
btn_tim2.lift()

# кнопки 3 этажа
btn_mov3 = Button(f3, text='Перейти на 1 этаж', command=lambda:raise_frame(f1))
btn_mov3.pack()
Label(f3, text='Этаж 3', font="Courier 30").pack()
btn_tim3 = Button(f3, text ="каб.Матем4", command=openMath4)
btn_tim3.place(x=125, y=350)
btn_tim3 = Button(f3, text ="каб.Англ1", command=openEngl1)
btn_tim3.place(x=950, y=500)
btn_tim3 = Button(f3, text ="каб.Англ2", command=openEngl2)
btn_tim3.place(x=650, y=500)
btn_tim3 = Button(f3, text ="Актовый зал")
btn_tim3.place(x=775, y=300)
btn_tim3 = Button(f3, text ="каб.Биологии1", command=openMath2)
btn_tim3.place(x=1500, y=350)
btn_tim3 = Button(f3, text ="каб.Химии", command=openChem)
btn_tim3.place(x=1100, y=550)
btn_tim3 = Button(f3, text ="каб.Биологии2", command=openBio1)
btn_tim3.place(x=1375, y=550)
btn_tim3 = Button(f3, text ="каб.Матем1", command=openMath1)
btn_tim3.place(x=550, y=550)
btn_tim3 = Button(f3, text ="каб.Матем2", command=openMath2)
btn_tim3.place(x=400, y=550)
btn_tim3 = Button(f3, text ="каб.Матем3", command=openMath3)
btn_tim3.place(x=275, y=550)
btn_tim3.columnconfigure(1, weight=1)
btn_tim3.rowconfigure(1, weight=1)
btn_tim3.lift()

raise_frame(f1)
root.mainloop()